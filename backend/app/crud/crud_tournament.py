# backend/app/crud/crud_tournament.py
from sqlalchemy.orm import Session, joinedload
from typing import Optional, List, Dict, Any
from sqlalchemy import func, desc, or_, case
from fastapi import HTTPException
from datetime import datetime
import math
import random

from app.models.models import Tournament, Match, Registration, Player, User, Payment, Court, MailCampaign, TournamentCategory
from app.schemas.tournament_schemas import TournamentCreate, TournamentUpdate, MatchScoreUpdate, MatchScheduleUpdate, GenerateDrawRequest, ManualMatchCreate, AdminMatchUpdate
from app.core.audit import log_action

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_tournament(db: Session, tournament: TournamentCreate):
    existing_slug = db.query(Tournament).filter(Tournament.slug == tournament.slug).first()
    if existing_slug:
        raise HTTPException(
            status_code=400,
            detail="Đường dẫn (Slug) này đã được sử dụng cho một giải đấu khác. Vui lòng chọn đường dẫn khác!"
        )

    db_tournament = Tournament(**tournament.model_dump())
    db.add(db_tournament)
    db.commit()
    db.refresh(db_tournament)
    return db_tournament
def delete_tournament_db(db: Session, tournament_id: int):
    """Xóa toàn b�" giải �ấu và các dữ li�!u liên quan (Cascade manual)"""
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="Không tìm thấy giải đấu để xóa.")

    # 1. Xóa các trận �ấu liên quan
    db.query(Match).filter(Match.tournament_id == tournament_id).delete()
    
    # 2. Xóa các lượt �Ēng ký
    db.query(Registration).filter(Registration.tournament_id == tournament_id).delete()
    
    # 3. Xóa các n�"i dung thi �ấu
    db.query(TournamentCategory).filter(TournamentCategory.tournament_id == tournament_id).delete()
    
    # 4. Xóa các chiến d�9ch email
    db.query(MailCampaign).filter(MailCampaign.tournament_id == tournament_id).delete()

    # 5. Cu�i cùng m�:i xóa giải �ấu
    db.delete(tournament)
    db.commit()
    return {"message": "Đã xóa giải đấu thành công!", "id": tournament_id}


def get_tournaments_with_counts(db: Session, skip: int = 0, limit: int = 10, status: str = None):
    query = db.query(Tournament).options(joinedload(Tournament.categories))
    if status:
        query = query.filter(Tournament.status == status)
    
    tournaments = query.order_by(
        case((Tournament.display_order > 0, 0), else_=1),
        Tournament.display_order.asc(),
        Tournament.start_date.desc(),
        Tournament.id.desc()
    ).offset(skip).limit(limit).all()
    
    # Tính s  slot  ã  Ēng ký cho từng giải
    for t in tournaments:
        t.current_participants = db.query(Registration).filter(
            Registration.tournament_id == t.id,
            Registration.status.in_(["confirmed", "pending"]),
            Registration.deleted_at.is_(None),
            Registration.is_locked == False
        ).count()
    return tournaments

def get_tournament_with_count(db: Session, tournament_id: int):
    t = db.query(Tournament).options(joinedload(Tournament.categories)).filter(Tournament.id == tournament_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Không tìm thấy giải đấu")
    
    t.current_participants = db.query(Registration).filter(
        Registration.tournament_id == t.id,
        Registration.status.in_(["confirmed", "pending"]),
        Registration.deleted_at.is_(None),
        Registration.is_locked == False
    ).count()
    return t

def get_system_stats(db: Session):
    total_tournaments = db.query(Tournament).count()
    total_registrations = db.query(Registration).count()
    active_tournaments = db.query(Tournament).filter(Tournament.status == "ongoing").count()
    pending_approvals = db.query(Registration).filter(Registration.status == "pending").count()
    
    revenue = db.query(func.sum(Payment.amount)).filter(Payment.status == "completed").scalar() or 0
    
    total_matches = db.query(Match).count()
    completed_matches = db.query(Match).filter(Match.status == "completed").count()
    
    completion_rate = 0
    if total_matches > 0:
        completion_rate = round((completed_matches / total_matches) * 100, 1)
        
    return {
        "total_tournaments": total_tournaments,
        "total_registrations": total_registrations,
        "active_tournaments": active_tournaments,
        "pending_approvals": pending_approvals,
        "revenue": float(revenue),
        "total_matches": total_matches,
        "completed_matches": completed_matches,
        "completion_rate": completion_rate
    }

def update_tournament_info(db: Session, tournament_id: int, tournament_in: TournamentCreate, admin_id: int):
    db_tour = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not db_tour:
        raise HTTPException(status_code=404, detail="Không tìm thấy giải đấu")

    for var, value in vars(tournament_in).items():
        setattr(db_tour, var, value)
        
    db.commit()
    db.refresh(db_tour)
    
    log_action(db, admin_id, "TOURNAMENT", "UPDATE", "Tournament", db_tour.id, None, {"name": db_tour.name}, "Cập nhật giải đấu")
    return db_tour
def generate_knockout_draw(db: Session, tournament_id: int, category_id: Optional[int] = None, draw_size: Optional[int] = None, round_names: Optional[List[str]] = None, draw_mode: str = "manual", representative_name: Optional[str] = None):
    query = db.query(Registration).filter(
        Registration.tournament_id == tournament_id,
        Registration.status.in_(["pending", "approved", "confirmed", "paid", "checked_in"]),
        Registration.deleted_at.is_(None),
        Registration.is_locked == False
    )
    if category_id:
        query = query.filter(Registration.tournament_category_id == category_id)
    regs = query.all()

    # In manual bracket mode, draw_size means participant/team count.
    # Example: 21 teams -> round 1 shows 11 bracket nodes: 10 full matches
    # plus 1 bye branch that admins can complete manually.
    participant_count = draw_size if draw_size and draw_size > 0 else len(regs)
    if participant_count <= 1:
        raise HTTPException(status_code=400, detail="Không đủ số đội để tạo nhánh đấu.")

    round_match_counts = []
    current_participants = participant_count
    while current_participants > 1:
        current_matches = math.ceil(current_participants / 2)
        round_match_counts.append(current_matches)
        current_participants = current_matches
    rounds_needed = len(round_match_counts)

    round_labels = {1: "FINAL", 2: "SF", 4: "QF", 8: "R16", 16: "R32", 32: "R64", 64: "R128"}

    match_del_query = db.query(Match).filter(Match.tournament_id == tournament_id, Match.stage_type == "knockout")
    if category_id:
        match_del_query = match_del_query.filter(Match.tournament_category_id == category_id)
    match_del_query.delete()
    db.flush()

    matches_by_round = {}
    for round_index, num_matches in enumerate(round_match_counts):
        matches_by_round[round_index] = []
        custom_label = round_names[round_index].strip() if round_names and round_index < len(round_names) and round_names[round_index] else None
        label = custom_label or round_labels.get(num_matches, f"R{num_matches * 2}")

        for i in range(num_matches):
            m = Match(
                tournament_id=tournament_id,
                tournament_category_id=category_id,
                stage_type="knockout",
                round_code=label,
                match_no=i + 1,
                status="pending",
                best_of_sets=3,
                elo_affected=True
            )
            db.add(m)
            matches_by_round[round_index].append(m)

    db.flush()

    # Link winners to the earliest available future match. This keeps the tree
    # valid even for odd participant counts where one side receives a bye.
    incoming_counts = {}
    for round_index in range(0, rounds_needed - 1):
        current_round = matches_by_round[round_index]
        for match in current_round:
            linked = False
            for future_round in range(round_index + 1, rounds_needed):
                for future_match in matches_by_round[future_round]:
                    current_incoming = incoming_counts.get(future_match.id, 0)
                    if current_incoming < 2:
                        match.next_match_id = future_match.id
                        incoming_counts[future_match.id] = current_incoming + 1
                        linked = True
                        break
                if linked:
                    break

    if draw_mode == "random":
        if not regs:
            raise HTTPException(status_code=400, detail="Không tìm thấy VĐV nào đăng ký hợp lệ để bốc thăm ngẫu nhiên.")
        
        # Xáo trộn ngẫu nhiên danh sách đăng ký
        random.shuffle(regs)
        count = len(regs)
        
        # Tạo slots tương ứng với vòng đầu tiên
        total_slots = 2 * round_match_counts[0]
        slots = [None] * total_slots
        for i in range(min(count, total_slots)):
            slots[i] = regs[i]
            
        slot_idx = 0
        for m in matches_by_round[0]:
            reg_a = slots[slot_idx] if slot_idx < len(slots) else None
            reg_b = slots[slot_idx + 1] if slot_idx + 1 < len(slots) else None
            slot_idx += 2
            
            m.side_a_registration_id = reg_a.id if reg_a else None
            m.side_b_registration_id = reg_b.id if reg_b else None
            
            # Tự động giải quyết các nhánh thắng miễn đấu (BYE)
            if reg_a and not reg_b:
                m.status = "completed"
                m.winner_side = "side_a"
                m.winner_registration_id = reg_a.id
                m.result_note = "BYE"
            elif not reg_a and reg_b:
                m.status = "completed"
                m.winner_side = "side_b"
                m.winner_registration_id = reg_b.id
                m.result_note = "BYE"
        
        db.flush()
        
        # Tiến cử các VĐV được miễn đấu vòng đầu (BYE) lên vòng 2
        for m in matches_by_round[0]:
            if m.status == "completed" and m.next_match_id and m.winner_registration_id:
                next_m = db.query(Match).filter(Match.id == m.next_match_id).first()
                if next_m:
                    if m.match_no % 2 != 0:
                        next_m.side_a_registration_id = m.winner_registration_id
                    else:
                        next_m.side_b_registration_id = m.winner_registration_id
                        
        # Ghi log audit
        if representative_name:
            log_action(db, None, "TOURNAMENT", "GENERATE", "Tournament", tournament_id, None, 
                       {"representative": representative_name, "mode": "random", "format": "knockout"}, 
                       f"Bốc thăm Knockout ngẫu nhiên bằng máy dưới sự đại diện của: {representative_name}")
    else:
        # Chế độ thủ công: Chỉ tạo các khung trống rỗng
        for m in matches_by_round[0]:
            m.side_a_registration_id = None
            m.side_b_registration_id = None

    db.commit()

    return {
        "message": "Bốc thăm ngẫu nhiên bằng máy thành công!" if draw_mode == "random" else "Đã tạo khung nhánh đấu thành công. Vui lòng tự ghép cặp thi đấu bằng tay.",
        "total_slots": participant_count,
        "first_round_matches": round_match_counts[0],
        "participant_count": participant_count,
        "rounds": rounds_needed
    }
def get_tournament_matches_detail(db: Session, tournament_id: int, category_id: Optional[int] = None):
    query = db.query(Match).filter(Match.tournament_id == tournament_id)
    if category_id:
        query = query.filter(Match.tournament_category_id == category_id)
    
    matches = query.order_by(Match.match_no).all()
    results = []
    for m in matches:
        p1_name = "Chua xac dinh"
        p2_name = "Chua xac dinh"
        p1_partner_name = None
        p2_partner_name = None
        p1_user_id = None
        p2_user_id = None
        p1_avatar = None
        p2_avatar = None
        p1_partner_user_id = None
        p1_partner_avatar = None
        p2_partner_user_id = None
        p2_partner_avatar = None
        
        reg_a = db.query(Registration).filter(Registration.id == m.side_a_registration_id).first() if m.side_a_registration_id else None
        if reg_a:
            p_a = db.query(Player).filter(Player.id == reg_a.player_id).first()
            if p_a:
                p1_user_id = p_a.user_id
                user_a = db.query(User).filter(User.id == p_a.user_id).first()
                if user_a:
                    p1_name = user_a.full_name
                    p1_avatar = user_a.avatar_url
            p1_partner_name = reg_a.partner_name
            p1_partner_user_id = reg_a.partner_user_id
            if p1_partner_user_id:
                user_partner_a = db.query(User).filter(User.id == p1_partner_user_id).first()
                if user_partner_a:
                    p1_partner_avatar = user_partner_a.avatar_url
                    p1_partner_name = user_partner_a.full_name # Override with user name if available

        reg_b = db.query(Registration).filter(Registration.id == m.side_b_registration_id).first() if m.side_b_registration_id else None
        if reg_b:
            p_b = db.query(Player).filter(Player.id == reg_b.player_id).first()
            if p_b:
                p2_user_id = p_b.user_id
                user_b = db.query(User).filter(User.id == p_b.user_id).first()
                if user_b:
                    p2_name = user_b.full_name
                    p2_avatar = user_b.avatar_url
            p2_partner_name = reg_b.partner_name
            p2_partner_user_id = reg_b.partner_user_id
            if p2_partner_user_id:
                user_partner_b = db.query(User).filter(User.id == p2_partner_user_id).first()
                if user_partner_b:
                    p2_partner_avatar = user_partner_b.avatar_url
                    p2_partner_name = user_partner_b.full_name # Override with user name if available
        
        court_name = db.query(Court.court_name).filter(Court.id == m.court_id).scalar() if m.court_id else None
                
        results.append({
            "id": m.id, "round_code": m.round_code, "match_no": m.match_no,
            "side_a_registration_id": m.side_a_registration_id,
            "side_b_registration_id": m.side_b_registration_id,
            "p1_name": p1_name, "p2_name": p2_name,
            "p1_avatar": p1_avatar, "p2_avatar": p2_avatar,
            "p1_partner_name": p1_partner_name, "p1_partner_user_id": p1_partner_user_id, "p1_partner_avatar": p1_partner_avatar,
            "p2_partner_name": p2_partner_name, "p2_partner_user_id": p2_partner_user_id, "p2_partner_avatar": p2_partner_avatar,
            "status": m.status,
            "court_id": m.court_id, 
            "court": court_name,
            "start_time": m.start_time.isoformat() if m.start_time else None, 
            "winner_side": m.winner_side,
            "score_a": m.set1_a,
            "score_b": m.set1_b,
            "set1_a": m.set1_a,
            "set1_b": m.set1_b,
            "set2_a": m.set2_a,
            "set2_b": m.set2_b,
            "set3_a": m.set3_a,
            "set3_b": m.set3_b,
            "tie_break_1_a": m.tie_break_1_a,
            "tie_break_1_b": m.tie_break_1_b,
            "tie_break_2_a": m.tie_break_2_a,
            "tie_break_2_b": m.tie_break_2_b,
            "tie_break_3_a": m.tie_break_3_a,
            "tie_break_3_b": m.tie_break_3_b,
            "p1_user_id": p1_user_id,
            "p2_user_id": p2_user_id,
            "referee_id": m.referee_id,
            "referee_name": m.referee_name,
            "referee_phone": m.referee_phone,
            "score_summary": getattr(m, 'score_summary', None) or getattr(m, 'result_note', None),
            "video_url": getattr(m, 'video_url', None),
            "live_stream_url": getattr(m, 'live_stream_url', None),
            "image_url": getattr(m, 'image_url', None),
            "advance_note": getattr(m, 'win_reason', None),
            "stage_type": m.stage_type,
            "group_id": m.group_id,
            "next_match_id": m.next_match_id,
            "tournament_category_id": m.tournament_category_id,
            "show_on_homepage": getattr(m, 'show_on_homepage', False)
        })
    return results

def _auto_link_manual_match(db: Session, match: Match):
    if match.stage_type not in ["knockout", "playoff"]:
        return

    base_query = db.query(Match).filter(
        Match.tournament_id == match.tournament_id,
        Match.stage_type == match.stage_type,
        Match.id != match.id
    )
    if match.tournament_category_id:
        base_query = base_query.filter(Match.tournament_category_id == match.tournament_category_id)

    all_matches = base_query.all()
    if not all_matches:
        return

    rounds_map = {}
    for item in all_matches + [match]:
        rounds_map.setdefault(item.round_code, []).append(item)

    ordered_rounds = sorted(
        rounds_map.items(),
        key=lambda entry: (-len(entry[1]), min((m.match_no or 0) for m in entry[1]))
    )
    round_codes = [round_code for round_code, _ in ordered_rounds]
    if match.round_code not in round_codes:
        return

    current_index = round_codes.index(match.round_code)
    desired_parent_match_no = math.ceil((match.match_no or 1) / 2)

    if current_index + 1 < len(round_codes):
        parent_round_code = round_codes[current_index + 1]
        parent_match = next(
            (
                candidate for candidate in rounds_map[parent_round_code]
                if (candidate.match_no or 0) == desired_parent_match_no
            ),
            None
        )
        if parent_match:
            match.next_match_id = parent_match.id

    if current_index > 0:
        child_round_code = round_codes[current_index - 1]
        candidate_children = sorted(
            [
                candidate for candidate in rounds_map[child_round_code]
                if math.ceil((candidate.match_no or 1) / 2) == (match.match_no or 1)
            ],
            key=lambda candidate: candidate.match_no or 0
        )
        incoming = db.query(func.count(Match.id)).filter(Match.next_match_id == match.id).scalar() or 0
        for child in candidate_children:
            if incoming >= 2:
                break
            if child.next_match_id in [None, match.next_match_id]:
                child.next_match_id = match.id
                incoming += 1

def validate_next_match_assignment(db: Session, match: Match, next_match_id: Optional[int]):
    if not next_match_id:
        return
    if next_match_id == match.id:
        raise HTTPException(status_code=400, detail="Không thể nối trận đấu đến chính nó.")

    next_match = db.query(Match).filter(Match.id == next_match_id).first()
    if not next_match:
        raise HTTPException(status_code=404, detail="Không tìm thấy trận đấu tiếp theo.")
    if next_match.tournament_id != match.tournament_id:
        raise HTTPException(status_code=400, detail="Trận đấu tiếp theo phải thuộc cùng giải đấu.")
    if match.tournament_category_id and next_match.tournament_category_id and next_match.tournament_category_id != match.tournament_category_id:
        raise HTTPException(status_code=400, detail="Trận đấu tiếp theo phải thuộc cùng nội dung.")

    visited = {match.id}
    cursor = next_match
    while cursor and cursor.next_match_id:
        if cursor.next_match_id in visited:
            raise HTTPException(status_code=400, detail="Liên kết nhánh đấu tạo vòng lặp.")
        visited.add(cursor.id)
        cursor = db.query(Match).filter(Match.id == cursor.next_match_id).first()

def _advance_winner_to_next_match(db: Session, match: Match, win_reg_id: Optional[int]):
    if not win_reg_id or not match.tournament_id or not match.next_match_id:
        return

    next_m = db.query(Match).filter(Match.id == match.next_match_id).first()
    if not next_m:
        return

    sibling_matches = db.query(Match).filter(
        Match.next_match_id == next_m.id,
        Match.tournament_id == match.tournament_id
    ).order_by(Match.match_no.asc(), Match.id.asc()).all()

    target_side = None
    for index, sibling in enumerate(sibling_matches):
        if sibling.id == match.id:
            target_side = "side_a" if index % 2 == 0 else "side_b"
            break

    if target_side == "side_a":
        next_m.side_a_registration_id = win_reg_id
    elif target_side == "side_b":
        next_m.side_b_registration_id = win_reg_id
    elif not next_m.side_a_registration_id:
        next_m.side_a_registration_id = win_reg_id
    elif not next_m.side_b_registration_id:
        next_m.side_b_registration_id = win_reg_id
    else:
        next_m.side_a_registration_id = win_reg_id

def create_manual_match_db(db: Session, tournament_id: int, payload: ManualMatchCreate):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="Không tìm thấy giải đấu")
    if payload.next_match_id:
        next_match = db.query(Match).filter(Match.id == payload.next_match_id).first()
        if not next_match:
            raise HTTPException(status_code=404, detail="Không tìm thấy trận đấu tiếp theo.")
        if next_match.tournament_id != tournament_id:
            raise HTTPException(status_code=400, detail="Trận đấu tiếp theo phải thuộc cùng giải đấu.")
        if payload.category_id and next_match.tournament_category_id and next_match.tournament_category_id != payload.category_id:
            raise HTTPException(status_code=400, detail="Trận đấu tiếp theo phải thuộc cùng nội dung.")

    category_id = payload.category_id
    round_code = (payload.round_code or "Vong moi").strip()
    stage_type = payload.stage_type or "knockout"
    max_no_query = db.query(func.max(Match.match_no)).filter(
        Match.tournament_id == tournament_id,
        Match.stage_type == stage_type,
        Match.round_code == round_code
    )
    if category_id:
        max_no_query = max_no_query.filter(Match.tournament_category_id == category_id)
    next_no = (max_no_query.scalar() or 0) + 1

    match = Match(
        tournament_id=tournament_id,
        tournament_category_id=category_id,
        stage_type=stage_type,
        round_code=round_code,
        match_no=payload.match_no or next_no,
        side_a_registration_id=payload.side_a_registration_id,
        side_b_registration_id=payload.side_b_registration_id,
        status=payload.status or "pending",
        court_id=payload.court_id,
        start_time=payload.start_time,
        referee_name=payload.referee_name or None,
        referee_phone=payload.referee_phone or None,
        live_stream_url=payload.live_stream_url or None,
        next_match_id=payload.next_match_id,
        best_of_sets=3,
        elo_affected=True,
    )
    db.add(match)
    db.flush()
    source_ids = payload.source_match_ids or []
    if source_ids:
        source_matches = db.query(Match).filter(
            Match.id.in_(source_ids),
            Match.tournament_id == tournament_id
        ).all()
        for source in source_matches:
            source.next_match_id = match.id
    if not payload.next_match_id and not source_ids:
        _auto_link_manual_match(db, match)
    db.commit()
    db.refresh(match)
    return {"message": "Đã thêm trận thủ công vào nhánh đấu", "id": match.id}

def update_match_admin_db(db: Session, match_id: int, payload: AdminMatchUpdate):
    match = db.query(Match).filter(Match.id == match_id).first()
    if not match:
        raise HTTPException(status_code=404, detail="Không tìm thấy trận đấu")

    data = payload.model_dump(exclude_unset=True)
    if data.get("side_a_registration_id") and data.get("side_a_registration_id") == data.get("side_b_registration_id"):
        raise HTTPException(status_code=400, detail="Không thể xếp cùng một VĐV/cặp đấu ở cả hai bên.")
    if "next_match_id" in data:
        validate_next_match_assignment(db, match, data["next_match_id"])

    for field in [
        "round_code", "match_no", "stage_type", "side_a_registration_id", "side_b_registration_id",
        "status", "court_id", "start_time", "referee_name", "referee_phone",
        "live_stream_url", "video_url", "image_url", "winner_side", "next_match_id", "show_on_homepage",
        "set1_a", "set1_b", "set2_a", "set2_b", "set3_a", "set3_b",
        "tie_break_1_a", "tie_break_1_b", "tie_break_2_a", "tie_break_2_b", "tie_break_3_a", "tie_break_3_b"
    ]:
        if field in data:
            setattr(match, field, data[field])

    if "advance_note" in data:
        match.win_reason = (data["advance_note"] or None)

    if "score" in data:
        match.result_note = data["score"]
        match.score_summary = data["score"]

    if data.get("winner_side"):
        match.winner_registration_id = match.side_a_registration_id if data["winner_side"] == "side_a" else match.side_b_registration_id
        _advance_winner_to_next_match(db, match, match.winner_registration_id)

    db.commit()
    return {"message": "Đã cập nhật thông tin điều hành trận đấu"}

def delete_match_from_draw_db(db: Session, match_id: int):
    match = db.query(Match).filter(Match.id == match_id).first()
    if not match:
        raise HTTPException(status_code=404, detail="Không tìm thấy trận đấu")

    db.query(Match).filter(Match.next_match_id == match_id).update(
        {Match.next_match_id: match.next_match_id},
        synchronize_session=False
    )
    db.delete(match)
    db.commit()
    return {"message": "Đã xóa khung trận đấu khỏi sơ đồ"}

def schedule_match_db(db: Session, match_id: int, payload: MatchScheduleUpdate):
    db_match = db.query(Match).filter(Match.id == match_id).first()
    if not db_match:
        raise HTTPException(status_code=404, detail="Không tìm thấy trận đấu")
    
    tournament = db.query(Tournament).filter(Tournament.id == db_match.tournament_id).first()
    if tournament:
        schedule_date = payload.start_time.date()
        if tournament.start_date and schedule_date < tournament.start_date:
            raise HTTPException(status_code=400, detail=f"Giải đấu bắt đầu từ ngày {tournament.start_date.strftime('%d/%m/%Y')}.")
        if tournament.end_date and schedule_date > tournament.end_date:
            raise HTTPException(status_code=400, detail=f"Giải đấu kết thúc vào ngày {tournament.end_date.strftime('%d/%m/%Y')}.")
            
    db_match.court_id = payload.court_id
    db_match.start_time = payload.start_time
    if payload.referee_id:
        db_match.referee_id = payload.referee_id
    db_match.referee_name = payload.referee_name
    db_match.referee_phone = payload.referee_phone
    db.commit()
    return {"message": "Đã cập nhật lịch thi đấu"}

def get_all_matches_detail(db: Session, limit: Optional[int] = None, show_on_homepage: Optional[bool] = None, status: Optional[str] = None):
    query = db.query(Match, Tournament, Court).outerjoin(
        Tournament, Match.tournament_id == Tournament.id
    ).outerjoin(
        Court, Match.court_id == Court.id
    )
    if show_on_homepage is not None:
        query = query.filter(Match.show_on_homepage == show_on_homepage)
    if status:
        status_list = [s.strip().lower() for s in status.split(",") if s.strip()]
        if status_list:
            query = query.filter(Match.status.in_(status_list))
    query = query.order_by(desc(Match.start_time))
    if limit:
        query = query.limit(limit)
    matches = query.all()

    # Helper lấy thông tin �ầy �ủ của team từ match và side
    def get_match_players_data(m, side):
        if side == "a":
            reg_id = m.side_a_registration_id
            p_id = m.player_a_id
            p2_id = m.player_a2_id
        else:
            reg_id = m.side_b_registration_id
            p_id = m.player_b_id
            p2_id = m.player_b2_id

        data = {"name": None, "avatar": None, "partner_name": None, "partner_avatar": None}

        # Cách 1: Qua registration (cho giải �ấu)
        if reg_id:
            reg = db.query(Registration).filter(Registration.id == reg_id).first()
            if reg:
                user = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == reg.player_id).first()
                data["name"] = user.full_name if user else None
                data["avatar"] = user.avatar_url if user else None
                data["partner_name"] = reg.partner_name
                
                if reg.partner_user_id:
                    p_user = db.query(User).filter(User.id == reg.partner_user_id).first()
                    if p_user:
                        data["partner_avatar"] = p_user.avatar_url
                elif getattr(reg, "partner_player_id", None):
                    p_user = db.query(User).join(Player).filter(Player.id == reg.partner_player_id).first()
                    if p_user:
                        data["partner_avatar"] = p_user.avatar_url
            return data

        # Cách 2: Qua direct player_id (cho trận giao hữu/thách �ấu)
        if p_id:
            user = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == p_id).first()
            if user:
                data["name"] = user.full_name
                data["avatar"] = user.avatar_url

        if p2_id:
            user2 = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == p2_id).first()
            if user2:
                data["partner_name"] = user2.full_name
                data["partner_avatar"] = user2.avatar_url

        return data

    results = []
    for m, t, c in matches:
        # Ưu tiên match_date của trận, fallback về start_time.date(), cu�i cùng là start_date giải
        if m.match_date:
            match_date = m.match_date
        elif m.start_time:
            match_date = m.start_time.date()
        else:
            match_date = t.start_date if t else None

        p1_data = get_match_players_data(m, "a")
        p2_data = get_match_players_data(m, "b")

        results.append({
            "id": m.id,
            "tournament_id": t.id if t else None,
            "tournament": t.name if t else "Giao hữu tự do",
            "tournament_start_date": t.start_date.isoformat() if t and t.start_date else None,
            "tournament_end_date": t.end_date.isoformat() if t and t.end_date else None,
            "location": (t.location if t else None) or "Saigontennistours Club",
            "round_code": m.round_code,
            "court": c.court_name if c else "Chua gan san",
            "date": match_date.isoformat() if match_date else None,
            "start_time": m.start_time.isoformat() if m.start_time else None,
            "start": m.start_time.strftime("%H:%M") if m.start_time else "--:--",
            "status": m.status,
            "p1_name": p1_data["name"],
            "p1_avatar": p1_data["avatar"],
            "p1_partner_name": p1_data["partner_name"],
            "p1_partner_avatar": p1_data["partner_avatar"],
            "p2_name": p2_data["name"],
            "p2_avatar": p2_data["avatar"],
            "p2_partner_name": p2_data["partner_name"],
            "p2_partner_avatar": p2_data["partner_avatar"],
            "winner_side": m.winner_side,
            "score": m.score_summary,
            "set1_a": m.set1_a,
            "set1_b": m.set1_b,
            "set2_a": m.set2_a,
            "set2_b": m.set2_b,
            "set3_a": m.set3_a,
            "set3_b": m.set3_b,
            "tie_break_1_a": m.tie_break_1_a,
            "tie_break_1_b": m.tie_break_1_b,
            "tie_break_2_a": m.tie_break_2_a,
            "tie_break_2_b": m.tie_break_2_b,
            "tie_break_3_a": m.tie_break_3_a,
            "tie_break_3_b": m.tie_break_3_b,
            "video_url": getattr(m, 'video_url', None),
            "show_on_homepage": getattr(m, 'show_on_homepage', False),
        })
    return results

def get_match_detail(db: Session, match_id: int):
    query = db.query(Match, Tournament, Court).outerjoin(
        Tournament, Match.tournament_id == Tournament.id
    ).outerjoin(
        Court, Match.court_id == Court.id
    ).filter(Match.id == match_id)
    
    row = query.first()
    if not row:
        return None
        
    m, t, c = row
    
    def get_match_players_data(m, side):
        if side == "a":
            reg_id = m.side_a_registration_id
            p_id = m.player_a_id
            p2_id = m.player_a2_id
        else:
            reg_id = m.side_b_registration_id
            p_id = m.player_b_id
            p2_id = m.player_b2_id

        data = {"name": None, "avatar": None, "partner_name": None, "partner_avatar": None}

        if reg_id:
            reg = db.query(Registration).filter(Registration.id == reg_id).first()
            if reg:
                user = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == reg.player_id).first()
                data["name"] = user.full_name if user else None
                data["avatar"] = user.avatar_url if user else None
                data["partner_name"] = reg.partner_name
                
                if reg.partner_user_id:
                    p_user = db.query(User).filter(User.id == reg.partner_user_id).first()
                    if p_user:
                        data["partner_avatar"] = p_user.avatar_url
                elif getattr(reg, "partner_player_id", None):
                    p_user = db.query(User).join(Player).filter(Player.id == reg.partner_player_id).first()
                    if p_user:
                        data["partner_avatar"] = p_user.avatar_url
            return data

        if p_id:
            user = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == p_id).first()
            if user:
                data["name"] = user.full_name
                data["avatar"] = user.avatar_url

        if p2_id:
            user2 = db.query(User).join(Player, Player.user_id == User.id).filter(Player.id == p2_id).first()
            if user2:
                data["partner_name"] = user2.full_name
                data["partner_avatar"] = user2.avatar_url

        return data

    if m.match_date:
        match_date = m.match_date
    elif m.start_time:
        match_date = m.start_time.date()
    else:
        match_date = t.start_date if t else None

    p1_data = get_match_players_data(m, "a")
    p2_data = get_match_players_data(m, "b")

    return {
        "id": m.id,
        "tournament_id": t.id if t else None,
        "tournament": t.name if t else "Giao hữu tự do",
        "tournament_start_date": t.start_date.isoformat() if t and t.start_date else None,
        "tournament_end_date": t.end_date.isoformat() if t and t.end_date else None,
        "location": (t.location if t else None) or "Saigontennistours Club",
        "round_code": m.round_code,
        "court": c.court_name if c else "Chua gan san",
        "court_id": m.court_id,
        "date": match_date.isoformat() if match_date else None,
        "start_time": m.start_time.isoformat() if m.start_time else None,
        "start": m.start_time.strftime("%H:%M") if m.start_time else "--:--",
        "status": m.status,
        "p1_name": p1_data["name"],
        "p1_avatar": p1_data["avatar"],
        "p1_partner_name": p1_data["partner_name"],
        "p1_partner_avatar": p1_data["partner_avatar"],
        "p2_name": p2_data["name"],
        "p2_avatar": p2_data["avatar"],
        "p2_partner_name": p2_data["partner_name"],
        "p2_partner_avatar": p2_data["partner_avatar"],
        "winner_side": m.winner_side,
        "score": m.score_summary,
        "set1_a": m.set1_a,
        "set1_b": m.set1_b,
        "set2_a": m.set2_a,
        "set2_b": m.set2_b,
        "set3_a": m.set3_a,
        "set3_b": m.set3_b,
        "tie_break_1_a": m.tie_break_1_a,
        "tie_break_1_b": m.tie_break_1_b,
        "tie_break_2_a": m.tie_break_2_a,
        "tie_break_2_b": m.tie_break_2_b,
        "tie_break_3_a": m.tie_break_3_a,
        "tie_break_3_b": m.tie_break_3_b,
        "best_of_sets": m.best_of_sets,
        "referee_name": m.referee_name,
        "referee_phone": m.referee_phone,
        "video_url": getattr(m, 'video_url', None),
        "show_on_homepage": getattr(m, 'show_on_homepage', False),
    }

def calculate_elo_and_update_match(db: Session, match_id: int, payload: MatchScoreUpdate):
    # 1. Tìm trận  ấu và kiỒm tra trạng thái
    match = db.query(Match).filter(Match.id == match_id).first()
    if not match or match.status == "completed":
        raise HTTPException(status_code=400, detail="Trận đấu không tồn tại hoặc đã kết thúc.")

    # 0. KiỒm tra xem trận  ấu có  ược phép tính ELO không
    if not getattr(match, 'elo_affected', False):
        # Nếu không tính ELO, chúng ta ch0 cập nhật trạng thái trận đấu
        match.status = "completed"
        match.score_summary = payload.score
        match.winner_side = payload.winner_side
        if payload.referee_id:
            match.referee_id = payload.referee_id
        match.referee_name = payload.referee_name
        match.referee_phone = payload.referee_phone
        
        # Cập nhật điểm set
        if payload.set1_a is not None: match.set1_a = payload.set1_a
        if payload.set1_b is not None: match.set1_b = payload.set1_b
        if payload.set2_a is not None: match.set2_a = payload.set2_a
        if payload.set2_b is not None: match.set2_b = payload.set2_b
        if payload.set3_a is not None: match.set3_a = payload.set3_a
        if payload.set3_b is not None: match.set3_b = payload.set3_b
        
        # Cập nhật điểm tie-break
        if payload.tie_break_1_a is not None: match.tie_break_1_a = payload.tie_break_1_a
        if payload.tie_break_1_b is not None: match.tie_break_1_b = payload.tie_break_1_b
        if payload.tie_break_2_a is not None: match.tie_break_2_a = payload.tie_break_2_a
        if payload.tie_break_2_b is not None: match.tie_break_2_b = payload.tie_break_2_b
        if payload.tie_break_3_a is not None: match.tie_break_3_a = payload.tie_break_3_a
        if payload.tie_break_3_b is not None: match.tie_break_3_b = payload.tie_break_3_b
        
        db.commit()
        return {"message": "Cập nhật tỷ số thành công (Không tính ELO)"}

    # 1. LẤY PLAYER ID CỦA 2 B�`N
    p1_id = None
    p2_id = None

    if match.tournament_id:
        # Trường hợp trận �ấu GIẢI: Lấy Player ID thông qua bảng Registration
        reg_a = db.query(Registration).filter(Registration.id == match.side_a_registration_id).first()
        reg_b = db.query(Registration).filter(Registration.id == match.side_b_registration_id).first()
        if reg_a: p1_id = reg_a.player_id
        if reg_b: p2_id = reg_b.player_id
        
        # Fallback: Nếu không tìm thấy qua Registration, lấy trực tiếp từ match (cho các trận tạo thủ công)
        if not p1_id: p1_id = match.player_a_id
        if not p2_id: p2_id = match.player_b_id
    else:
        # Trường hợp trận GIAO HỮU: Lấy trực tiếp từ player_a_id và player_b_id
        p1_id = match.player_a_id
        p2_id = match.player_b_id

    # 3. KiỒm tra tính �ầy �ủ của 2 vận ��"ng viên
    if not p1_id or not p2_id:
        # Nếu thiếu 1 bên (lẻ ��"i, thắng bye), ta ch�0 cập nhật kết quả mà không tính ELO
        match.status = "completed"
        match.winner_side = payload.winner_side
        win_reg_id = match.side_a_registration_id if payload.winner_side == "side_a" else match.side_b_registration_id
        match.winner_registration_id = win_reg_id
        match.result_note = payload.score
        if payload.video_url is not None:
            match.video_url = payload.video_url
        if payload.image_url is not None:
            match.image_url = payload.image_url
        if payload.referee_id:
            match.referee_id = payload.referee_id
        match.referee_name = payload.referee_name
        match.referee_phone = payload.referee_phone
        if payload.set1_a is not None: match.set1_a = payload.set1_a
        if payload.set1_b is not None: match.set1_b = payload.set1_b
        if payload.set2_a is not None: match.set2_a = payload.set2_a
        if payload.set2_b is not None: match.set2_b = payload.set2_b
        if payload.set3_a is not None: match.set3_a = payload.set3_a
        if payload.set3_b is not None: match.set3_b = payload.set3_b
        if payload.tie_break_1_a is not None: match.tie_break_1_a = payload.tie_break_1_a
        if payload.tie_break_1_b is not None: match.tie_break_1_b = payload.tie_break_1_b
        if payload.tie_break_2_a is not None: match.tie_break_2_a = payload.tie_break_2_a
        if payload.tie_break_2_b is not None: match.tie_break_2_b = payload.tie_break_2_b
        if payload.tie_break_3_a is not None: match.tie_break_3_a = payload.tie_break_3_a
        if payload.tie_break_3_b is not None: match.tie_break_3_b = payload.tie_break_3_b
        
        _advance_winner_to_next_match(db, match, win_reg_id)
        
        db.commit()
        return {"message": "Cập nhật tỷ số thành công cho trận đấu lẻ (Thắng bye/Lẻ đội)!"}


    # 4. Xác ��9nh ai thắng ai thua dựa trên payload gửi lên
    win_p_id = p1_id if payload.winner_side == "side_a" else p2_id
    lose_p_id = p2_id if payload.winner_side == "side_a" else p1_id
    
    # Xác ��9nh registration_id của người thắng (ch�0 dùng cho logic tiến vào vòng sau của Giải)
    win_reg_id = match.side_a_registration_id if payload.winner_side == "side_a" else match.side_b_registration_id

    winner_p = db.query(Player).filter(Player.id == win_p_id).first()
    loser_p = db.query(Player).filter(Player.id == lose_p_id).first()

    if not winner_p or not loser_p:
        raise HTTPException(status_code=404, detail="Không tìm thấy hồ sơ vận động viên.")

    # 5. THUẬT TOÁN ELO
    # ELO points are now managed manually by admins.
    # Keep the old automatic formula here for future reuse if needed:
    #
    # K = 32
    # Ra = winner_p.elo_points
    # Rb = loser_p.elo_points
    # E_winner = 1 / (1 + 10 ** ((Rb - Ra) / 400))
    # elo_gain = round(K * (1 - E_winner))
    
    # Cập nhật ch�0 s� cho người thắng (và ��ng ��"i nếu có)
    def update_p_stats(p, is_win):
        if is_win:
            p.wins += 1
        else:
            p.losses += 1
        p.matches_played += 1

    update_p_stats(winner_p, True)
    update_p_stats(loser_p, False)

    # Nếu là �ánh �ôi, cập nhật cho cả ��ng ��"i
    if match.tournament_id:
        reg_win = db.query(Registration).filter(Registration.id == win_reg_id).first()
        lose_reg_id = match.side_b_registration_id if payload.winner_side == "side_a" else match.side_a_registration_id
        reg_lose = db.query(Registration).filter(Registration.id == lose_reg_id).first()

        if reg_win and reg_win.partner_player_id:
            partner_win = db.query(Player).filter(Player.id == reg_win.partner_player_id).first()
            if partner_win:
                update_p_stats(partner_win, True)

        if reg_lose and reg_lose.partner_player_id:
            partner_lose = db.query(Player).filter(Player.id == reg_lose.partner_player_id).first()
            if partner_lose:
                update_p_stats(partner_lose, False)

    # 6. Cập nhật thông tin trận �ấu
    match.status = "completed"
    match.winner_side = payload.winner_side
    match.winner_registration_id = win_reg_id # Lưu reg_id nếu có
    match.result_note = payload.score
    if payload.video_url is not None:
        match.video_url = payload.video_url
    if payload.image_url is not None:
        match.image_url = payload.image_url
    if payload.referee_id:
        match.referee_id = payload.referee_id
    match.referee_name = payload.referee_name
    match.referee_phone = payload.referee_phone
    if payload.set1_a is not None: match.set1_a = payload.set1_a
    if payload.set1_b is not None: match.set1_b = payload.set1_b
    if payload.set2_a is not None: match.set2_a = payload.set2_a
    if payload.set2_b is not None: match.set2_b = payload.set2_b
    if payload.set3_a is not None: match.set3_a = payload.set3_a
    if payload.set3_b is not None: match.set3_b = payload.set3_b
    if payload.tie_break_1_a is not None: match.tie_break_1_a = payload.tie_break_1_a
    if payload.tie_break_1_b is not None: match.tie_break_1_b = payload.tie_break_1_b
    if payload.tie_break_2_a is not None: match.tie_break_2_a = payload.tie_break_2_a
    if payload.tie_break_2_b is not None: match.tie_break_2_b = payload.tie_break_2_b
    if payload.tie_break_3_a is not None: match.tie_break_3_a = payload.tie_break_3_a
    if payload.tie_break_3_b is not None: match.tie_break_3_b = payload.tie_break_3_b

    # 7. Xử lý logic thĒng hạng nếu là trận �ấu giải[cite: 33]
    message_suffix = ""
    if match.tournament_id:
        # KiỒm tra xem còn trận �ấu nào chưa xong không
        remaining_matches = db.query(Match).filter(
            Match.tournament_id == match.tournament_id,
            Match.id != match.id, # Trừ trận hi�!n tại vừa xong
            Match.status.in_(["pending", "scheduled", "ongoing"])
        ).count()

        if remaining_matches == 0:
            tournament = db.query(Tournament).filter(Tournament.id == match.tournament_id).first()
            if tournament and tournament.status != "finished":
                tournament.status = "finished"
                # Cập nhật ID nhà vô ��9ch nếu là trận Chung kết
                if match.round_code in ["FINAL", "F"] and hasattr(tournament, 'winner_player_id'):
                    tournament.winner_player_id = winner_p.id
                message_suffix = f" Giải đấu đã chính thức khép lại. Chúc mừng {winner_p.full_name if hasattr(winner_p, 'full_name') else winner_p.id}!"
        else:
            _advance_winner_to_next_match(db, match, win_reg_id)

    db.commit()
    return {"message": f"Cập nhật kết quả thành công! {message_suffix}"}

def get_public_bracket_detail(db: Session, tournament_id: int, category_id: Optional[int] = None):
    query = db.query(Match).filter(Match.tournament_id == tournament_id)
    if category_id:
        query = query.filter(Match.tournament_category_id == category_id)
    matches = query.all()
    
    def get_player_data(reg_id):
        if not reg_id: return {"name": "Chua xac dinh", "user_id": None, "avatar_url": None, "partner_name": None, "partner_user_id": None, "partner_avatar_url": None}
        reg = db.query(Registration).filter(Registration.id == reg_id).first()
        if not reg: return {"name": "Chua xac dinh", "user_id": None, "avatar_url": None, "partner_name": None, "partner_user_id": None, "partner_avatar_url": None}
        
        user = db.query(User).join(Player).filter(Player.id == reg.player_id).first()
        data = {
            "name": user.full_name if user else "Chua xac dinh",
            "user_id": user.id if user else None,
            "avatar_url": user.avatar_url if user else None,
            "partner_name": None,
            "partner_user_id": None
        }
        
        if reg.partner_user_id:
            partner_user = db.query(User).filter(User.id == reg.partner_user_id).first()
            if partner_user:
                data["partner_name"] = partner_user.full_name
                data["partner_user_id"] = partner_user.id
                data["partner_avatar_url"] = partner_user.avatar_url
        elif getattr(reg, "partner_player_id", None):
            partner_user = db.query(User).join(Player).filter(Player.id == reg.partner_player_id).first()
            if partner_user:
                data["partner_name"] = partner_user.full_name
                data["partner_user_id"] = partner_user.id
                data["partner_avatar_url"] = partner_user.avatar_url
        elif reg.partner_name: 
            data["partner_name"] = reg.partner_name
                
        return data

    results = []
    for m in matches:
        p1_data = get_player_data(m.side_a_registration_id)
        p2_data = get_player_data(m.side_b_registration_id)
        
        court_name = db.query(Court.court_name).filter(Court.id == m.court_id).scalar() if m.court_id else None

        results.append({
            "id": m.id, "match_no": m.match_no, "round_code": m.round_code,
            "category_id": m.tournament_category_id,
            "p1_name": p1_data["name"],
            "p1_user_id": p1_data["user_id"],
            "p1_avatar": p1_data["avatar_url"],
            "p1_partner_name": p1_data["partner_name"],
            "p1_partner_user_id": p1_data["partner_user_id"],
            "p1_partner_avatar": p1_data["partner_avatar_url"],
            
            "p2_name": p2_data["name"],
            "p2_user_id": p2_data["user_id"],
            "p2_avatar": p2_data["avatar_url"],
            "p2_partner_name": p2_data["partner_name"],
            "p2_partner_user_id": p2_data["partner_user_id"],
            "p2_partner_avatar": p2_data["partner_avatar_url"],
            
            "winner_side": m.winner_side, "status": m.status,
            "start_time": m.start_time,
            "match_date": m.match_date,
            "score": m.result_note,
            "score_summary": m.score_summary,
            "score_a": m.set1_a,
            "score_b": m.set1_b,
            "court_id": m.court_id,
            "court": court_name,
            "live_stream_url": getattr(m, "live_stream_url", None),
            "video_url": getattr(m, "video_url", None),
            "image_url": getattr(m, "image_url", None),
            "advance_note": getattr(m, "win_reason", None),
            "referee_name": m.referee_name or (db.query(User.full_name).filter(User.id == m.referee_id).scalar() if m.referee_id else None),
            "referee_phone": m.referee_phone
        })
    return results

def export_tournament_data_to_excel(db: Session, tournament_id: int):
    # 1. Lấy thông tin giải �ấu
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="Không tìm thấy giải đấu")

    # Tạo Workbook Excel m�:i
    wb = Workbook()
    
    # SHEET 1: DANH SÁCH VẬN Đ��NG VI�`N Đ�NG KÝ
    ws_players = wb.active
    ws_players.title = "Danh sách VĐV"
    
    # Style cho tiêu �ề
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["STT", "Họ Tên", "Số điện thoại", "Email", "Trình độ", "Trạng thái", "Thanh toán"]
    ws_players.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws_players.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Set ��" r�"ng c�"t
    ws_players.column_dimensions['B'].width = 25
    ws_players.column_dimensions['C'].width = 15
    ws_players.column_dimensions['D'].width = 25
    ws_players.column_dimensions['E'].width = 15
    ws_players.column_dimensions['F'].width = 15
    ws_players.column_dimensions['G'].width = 15

    # Lấy dữ liệu VĐV
    regs = db.query(Registration, Player, User).join(
        Player, Registration.player_id == Player.id
    ).join(
        User, Player.user_id == User.id
    ).filter(
        Registration.tournament_id == tournament_id,
        Registration.deleted_at.is_(None),
        Registration.is_locked == False
    ).all()

    for idx, (reg, player, user) in enumerate(regs, start=1):
        ws_players.append([
            idx,
            user.full_name,
            user.phone,
            user.email,
            player.skill_level or "N/A",
            "Đã Check-in" if reg.status == "checked_in" else "Đã duyệt" if reg.status == "confirmed" else "Chờ duyệt",
            "Đã thanh toán" if reg.payment_status == "paid" else "Chưa thanh toán"
        ])

    # SHEET 2: KẾT QUẢ TRẬN ĐẤU (BRACKET)

    ws_matches = wb.create_sheet(title="Kết quả Thi đấu")
    matches_headers = ["Trận số", "Vòng đấu", "VĐV A", "VĐV B", "Tỷ số", "Người thắng", "Trạng thái"]
    ws_matches.append(matches_headers)

    for col in range(1, len(matches_headers) + 1):
        cell = ws_matches.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws_matches.column_dimensions['C'].width = 25
    ws_matches.column_dimensions['D'].width = 25
    ws_matches.column_dimensions['E'].width = 20
    ws_matches.column_dimensions['F'].width = 25

    # Lấy dữ li�!u Match
    matches = db.query(Match).filter(Match.tournament_id == tournament_id).order_by(Match.match_no).all()
    
    def get_player_name(reg_id):
        if not reg_id: return "Chờ xếp nhánh"
        r = db.query(Registration).filter(Registration.id == reg_id).first()
        if not r: return "N/A"
        u = db.query(User).join(Player).filter(Player.id == r.player_id).first()
        if not u: return "VĐV"
        return f"{u.full_name} - {r.partner_name}" if r.partner_name else u.full_name

    for m in matches:
        p1_name = get_player_name(m.side_a_registration_id)
        p2_name = get_player_name(m.side_b_registration_id)
        
        winner_name = ""
        if m.status == "completed":
            winner_name = p1_name if m.winner_side == "side_a" else p2_name

        ws_matches.append([
            m.match_no,
            m.round_code,
            p1_name,
            p2_name,
            m.result_note or "-",
            winner_name,
            "Đã xong" if m.status == "completed" else "Chưa đấu"
        ])

    # Lưu vào b�" nh�: tạm (BytesIO) �Ồ gửi thẳng về Frontend mà không cần lưu rác trong �" cứng máy chủ
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Trả về file stream và tên file an toàn
    safe_name = "".join([c if c.isalnum() else "_" for c in tournament.name])
    file_name = f"BaoCao_{safe_name}.xlsx"
    
    return stream, file_name

def get_tournament_and_valid_emails(db: Session, tournament_id: int):
    # Lấy thông tin giải �ấu
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        return None, []

    # Lấy danh sách email hợp l�!
    valid_regs = db.query(User.email).join(
        Player, User.id == Player.user_id
    ).join(
        Registration, Player.id == Registration.player_id
    ).filter(
        Registration.tournament_id == tournament_id, 
        Registration.deleted_at.is_(None)
    ).all()
    
    # Lọc bỏ các phần tử r�ng
    bcc_emails = [reg[0] for reg in valid_regs if reg[0]]
    
    return tournament, bcc_emails

def save_mail_campaign(
    db: Session, 
    tournament_id: int, 
    subject: str, 
    message: str, 
    total_recipients: int,
    scheduled_at = None,   # <--- Thêm dòng này
    status: str = "pending" # <--- Thêm dòng này
):
    new_campaign = MailCampaign(
        tournament_id=tournament_id,
        subject=subject,
        message=message,
        total_recipients=total_recipients,
        scheduled_at=scheduled_at, # <--- Lưu vào DB
        status=status              # <--- Lưu vào DB
    )
    db.add(new_campaign)
    db.commit()
    db.refresh(new_campaign)
    return new_campaign

def generate_round_robin_draw(db: Session, tournament_id: int, category_id: int, num_groups: int = 1, draw_size: Optional[int] = None, draw_mode: str = "manual", representative_name: Optional[str] = None):
    """Generate round-robin group matches.

    When draw_size is provided, it is treated as the number of teams/pairs in
    the same manual setup style as knockout. Missing registrations are kept as
    empty slots so admins can assign them later.
    """
    if num_groups < 1:
        num_groups = 1

    query = db.query(Registration).filter(
        Registration.tournament_id == tournament_id,
        Registration.status.in_(["pending", "approved", "confirmed", "paid", "checked_in"]),
        Registration.deleted_at.is_(None),
        Registration.is_locked == False
    )
    if category_id:
        query = query.filter(Registration.tournament_category_id == category_id)
    players = query.all()

    if draw_mode == "random":
        if not players:
            raise HTTPException(status_code=400, detail="Không tìm thấy VĐV nào đăng ký hợp lệ để bốc thăm ngẫu nhiên.")
        # Xáo trộn ngẫu nhiên VĐV trước khi xếp bảng
        random.shuffle(players)
        participant_count = len(players)
        slots = [{"registration": p, "is_bye": False} for p in players]
    else:
        participant_count = draw_size if draw_size and draw_size > 0 else len(players)
        if participant_count < 2:
            raise HTTPException(status_code=400, detail="Không đủ số đội để tạo lịch thi đấu vòng tròn.")
        slots = [{"registration": None, "is_bye": False} for _ in range(participant_count)]

    match_del_query = db.query(Match).filter(
        Match.tournament_id == tournament_id,
        Match.stage_type == "group_stage"
    )
    if category_id:
        match_del_query = match_del_query.filter(Match.tournament_category_id == category_id)
    match_del_query.delete()
    db.flush()

    groups = [slots[i::num_groups] for i in range(num_groups)]

    match_no = 1
    for group_idx, group_players in enumerate(groups):
        group_id = group_idx + 1
        n = len(group_players)

        if n % 2 != 0:
            group_players.append({"registration": None, "is_bye": True})
            n += 1

        for round_num in range(n - 1):
            for i in range(n // 2):
                p1 = group_players[i]
                p2 = group_players[n - 1 - i]

                if p1.get("is_bye") or p2.get("is_bye"):
                    continue

                reg_a = p1.get("registration")
                reg_b = p2.get("registration")
                new_match = Match(
                    tournament_id=tournament_id,
                    tournament_category_id=category_id,
                    stage_type="group_stage",
                    group_id=group_id,
                    round_code=f"G{group_id}-R{round_num + 1}",
                    match_no=match_no,
                    side_a_registration_id=reg_a.id if reg_a else None,
                    side_b_registration_id=reg_b.id if reg_b else None,
                    status="pending",
                    best_of_sets=3,
                    elo_affected=True
                )
                db.add(new_match)
                match_no += 1

            group_players.insert(1, group_players.pop())

    # Ghi log audit
    if draw_mode == "random" and representative_name:
        log_action(db, None, "TOURNAMENT", "GENERATE", "Tournament", tournament_id, None, 
                   {"representative": representative_name, "mode": "random", "format": "round_robin"}, 
                   f"Bốc thăm Vòng bảng ngẫu nhiên bằng máy dưới sự đại diện của: {representative_name}")

    db.commit()
    return {
        "message": "Bốc thăm ngẫu nhiên bằng máy thành công!" if draw_mode == "random" else "Da tao lich thi dau vong tron thanh cong.",
        "total_slots": participant_count,
        "num_groups": num_groups,
        "matches_created": match_no - 1
    }

def parse_score_string(score_str: str):
    # Tra ve tuple (sets_a, sets_b, games_a, games_b)
    if not score_str:
        return 0, 0, 0, 0
    score_str = score_str.strip()
    if not score_str:
        return 0, 0, 0, 0
    sets_a = 0
    sets_b = 0
    games_a = 0
    games_b = 0
    set_tokens = score_str.split()
    for token in set_tokens:
        if "-" not in token:
            continue
        try:
            parts = token.split("-")
            if len(parts) != 2:
                continue
            str_a = parts[0].split("(")[0].strip()
            str_b = parts[1].split("(")[0].strip()
            val_a = int(str_a)
            val_b = int(str_b)
            games_a += val_a
            games_b += val_b
            if val_a > val_b:
                sets_a += 1
            elif val_b > val_a:
                sets_b += 1
        except Exception:
            continue
    return sets_a, sets_b, games_a, games_b

def calculate_tournament_standings(db: Session, tournament_id: int, category_id: Optional[int] = None):
    """Hàm lõi tính �iỒm (Dùng cho cả Vòng tròn và Xếp hạng t�"ng thỒ)"""
    # 1. Thử lấy các trận vòng bảng trư�:c
    query = db.query(Match).filter(
        Match.tournament_id == tournament_id,
        Match.stage_type == "group_stage",
        Match.status == "completed" 
    )
    if category_id:
        query = query.filter(Match.tournament_category_id == category_id)
    
    matches = query.all()

    # 2. Nếu không có trận vòng bảng nào, lấy tất cả các trận �ã xong của giải (cho Knockout/Playoff)
    if not matches:
        query = db.query(Match).filter(
            Match.tournament_id == tournament_id,
            Match.status == "completed"
        )
        if category_id:
            query = query.filter(Match.tournament_category_id == category_id)
        matches = query.all()

    standings = {}

    def safe_int(val):
        return int(val) if val is not None else 0

    for match in matches:
        # Xác ��9nh tên bảng
        if match.stage_type == "group_stage" and match.group_id:
            group = f"Bảng {match.group_id}"
        else:
            group = "Xếp hạng tổng thể"

        if group not in standings:
            standings[group] = {}

        p1_id = match.side_a_registration_id
        p2_id = match.side_b_registration_id

        if not p1_id or not p2_id:
            continue

        for p_id in [p1_id, p2_id]:
            if p_id not in standings[group]:
                user_record = db.query(
                    User.full_name, 
                    User.avatar_url,
                    Registration.partner_name, 
                    Registration.partner_player_id, 
                    Registration.partner_user_id,
                    Registration.player_id, 
                    User.id
                ).join(
                    Player, User.id == Player.user_id
                ).join(
                    Registration, Player.id == Registration.player_id
                ).filter(
                    Registration.id == p_id
                ).first()
                
                player_name = user_record[0] if user_record else "Unknown"
                player_avatar = user_record[1] if user_record else None
                partner_name = user_record[2] if user_record else None
                partner_player_id = user_record[3] if user_record else None
                partner_user_id_reg = user_record[4] if user_record else None
                player_user_id = user_record[6] if user_record else None
                
                # Nếu có mapping ID ��ng ��"i, lấy User ID của ��ng ��"i
                partner_user_id = partner_user_id_reg
                partner_avatar = None
                
                if partner_user_id:
                    p_user = db.query(User).filter(User.id == partner_user_id).first()
                    if p_user:
                        partner_name = p_user.full_name
                        partner_avatar = p_user.avatar_url
                elif partner_player_id:
                    p_user = db.query(User).join(Player).filter(Player.id == partner_player_id).first()
                    if p_user:
                        partner_name = p_user.full_name
                        partner_user_id = p_user.id
                        partner_avatar = p_user.avatar_url
                    
                standings[group][p_id] = {
                    "player_name": player_name, 
                    "player_avatar": player_avatar,
                    "player_id": player_user_id,
                    "partner_name": partner_name,
                    "partner_avatar": partner_avatar,
                    "partner_player_id": partner_user_id,
                    "played": 0, "won": 0, "lost": 0, "points": 0,
                    "sets_won": 0, "sets_lost": 0, 
                    "games_won": 0, "games_lost": 0,
                    "tb_won": 0, "tb_lost": 0
                }

        p1_games = safe_int(match.set1_a) + safe_int(match.set2_a) + safe_int(match.set3_a)
        p2_games = safe_int(match.set1_b) + safe_int(match.set2_b) + safe_int(match.set3_b)

        p1_sets = 0
        p2_sets = 0
        if safe_int(match.set1_a) > safe_int(match.set1_b): p1_sets += 1
        elif safe_int(match.set1_b) > safe_int(match.set1_a): p2_sets += 1
        if safe_int(match.set2_a) > safe_int(match.set2_b): p1_sets += 1
        elif safe_int(match.set2_b) > safe_int(match.set2_a): p2_sets += 1
        if safe_int(match.set3_a) > safe_int(match.set3_b): p1_sets += 1
        elif safe_int(match.set3_b) > safe_int(match.set3_a): p2_sets += 1

        # Neu khong co diem set/game nao o database phi chuan (bi NULL), tu dong parse tu chuoi ty so
        if p1_games == 0 and p2_games == 0 and p1_sets == 0 and p2_sets == 0:
            score_str = match.score_summary or match.result_note
            p1_sets, p2_sets, p1_games, p2_games = parse_score_string(score_str)

        # Tính toán điểm tie-break
        p1_tb = safe_int(match.tie_break_1_a) + safe_int(match.tie_break_2_a) + safe_int(match.tie_break_3_a)
        p2_tb = safe_int(match.tie_break_1_b) + safe_int(match.tie_break_2_b) + safe_int(match.tie_break_3_b)

        is_p1_winner = match.winner_registration_id == p1_id
        standings[group][p1_id]["played"] += 1
        standings[group][p1_id]["won"] += 1 if is_p1_winner else 0
        standings[group][p1_id]["lost"] += 0 if is_p1_winner else 1
        standings[group][p1_id]["points"] += 3 if is_p1_winner else 0
        standings[group][p1_id]["sets_won"] += p1_sets
        standings[group][p1_id]["sets_lost"] += p2_sets
        standings[group][p1_id]["games_won"] += p1_games
        standings[group][p1_id]["games_lost"] += p2_games
        standings[group][p1_id]["tb_won"] += p1_tb
        standings[group][p1_id]["tb_lost"] += p2_tb

        is_p2_winner = match.winner_registration_id == p2_id
        standings[group][p2_id]["played"] += 1
        standings[group][p2_id]["won"] += 1 if is_p2_winner else 0
        standings[group][p2_id]["lost"] += 0 if is_p2_winner else 1
        standings[group][p2_id]["points"] += 3 if is_p2_winner else 0
        standings[group][p2_id]["sets_won"] += p2_sets
        standings[group][p2_id]["sets_lost"] += p1_sets
        standings[group][p2_id]["games_won"] += p2_games
        standings[group][p2_id]["games_lost"] += p1_games
        standings[group][p2_id]["tb_won"] += p2_tb
        standings[group][p2_id]["tb_lost"] += p1_tb

    result = []
    for group_name, players in standings.items():
        for p_id, stats in players.items():
            stats["set_diff"] = stats["sets_won"] - stats["sets_lost"]
            stats["game_diff"] = stats["games_won"] - stats["games_lost"]
            stats["tb_diff"] = stats["tb_won"] - stats["tb_lost"]

        sorted_players = sorted(
            players.items(), 
            key=lambda x: (x[1]['points'], x[1]['set_diff'], x[1]['game_diff'], x[1]['tb_diff']), 
            reverse=True
        )
        
        result.append({
            "group_name": group_name,
            "rankings": [{"registration_id": k, **v} for k, v in sorted_players]
        })

    return result

def generate_playoff_draw(db: Session, tournament_id: int, category_id: int, advancers_per_group: int = 2):
    """Finalize group standings and create a knockout playoff tree."""
    if advancers_per_group < 1:
        raise ValueError("So VDV di tiep moi bang phai lon hon 0.")

    all_group_rows = db.query(Match.group_id).filter(
        Match.tournament_id == tournament_id,
        Match.tournament_category_id == category_id,
        Match.stage_type == "group_stage",
        Match.group_id.isnot(None)
    ).distinct().all()
    expected_group_ids = sorted(int(row[0]) for row in all_group_rows if row[0] is not None)

    if not expected_group_ids:
        raise ValueError("Chua co bang dau nao de chot Playoff.")

    group_matches_query = db.query(Match).filter(
        Match.tournament_id == tournament_id,
        Match.tournament_category_id == category_id,
        Match.stage_type == "group_stage",
        or_(
            Match.status == "completed",
            Match.winner_registration_id.isnot(None)
        ),
        Match.group_id.isnot(None)
    )
    group_matches = group_matches_query.order_by(Match.group_id.asc(), Match.round_code.asc(), Match.match_no.asc()).all()

    if not group_matches:
        raise ValueError("Chua co tran vong bang nao hoan thanh. Vui long cap nhat ty so.")

    grouped_stats = {}

    def safe_int(val):
        return int(val) if val is not None else 0

    def ensure_stat(group_id: int, registration_id: int):
        if group_id not in grouped_stats:
            grouped_stats[group_id] = {}
        if registration_id not in grouped_stats[group_id]:
            grouped_stats[group_id][registration_id] = {
                "registration_id": registration_id,
                "played": 0,
                "won": 0,
                "lost": 0,
                "points": 0,
                "sets_won": 0,
                "sets_lost": 0,
                "games_won": 0,
                "games_lost": 0,
            }
        return grouped_stats[group_id][registration_id]

    for match in group_matches:
        if not match.side_a_registration_id or not match.side_b_registration_id or not match.winner_registration_id:
            continue

        group_id = int(match.group_id)
        side_a = ensure_stat(group_id, match.side_a_registration_id)
        side_b = ensure_stat(group_id, match.side_b_registration_id)

        a_games = safe_int(match.set1_a) + safe_int(match.set2_a) + safe_int(match.set3_a)
        b_games = safe_int(match.set1_b) + safe_int(match.set2_b) + safe_int(match.set3_b)

        a_sets = 0
        b_sets = 0
        for a_score, b_score in [
            (match.set1_a, match.set1_b),
            (match.set2_a, match.set2_b),
            (match.set3_a, match.set3_b),
        ]:
            a_score = safe_int(a_score)
            b_score = safe_int(b_score)
            if a_score > b_score:
                a_sets += 1
            elif b_score > a_score:
                b_sets += 1

        a_won = match.winner_registration_id == match.side_a_registration_id
        b_won = match.winner_registration_id == match.side_b_registration_id

        side_a["played"] += 1
        side_a["won"] += 1 if a_won else 0
        side_a["lost"] += 0 if a_won else 1
        side_a["points"] += 3 if a_won else 0
        side_a["sets_won"] += a_sets
        side_a["sets_lost"] += b_sets
        side_a["games_won"] += a_games
        side_a["games_lost"] += b_games

        side_b["played"] += 1
        side_b["won"] += 1 if b_won else 0
        side_b["lost"] += 0 if b_won else 1
        side_b["points"] += 3 if b_won else 0
        side_b["sets_won"] += b_sets
        side_b["sets_lost"] += a_sets
        side_b["games_won"] += b_games
        side_b["games_lost"] += a_games

    group_tops = []
    insufficient_groups = []
    for group_id in expected_group_ids:
        players = list(grouped_stats.get(group_id, {}).values())
        for player in players:
            player["set_diff"] = player["sets_won"] - player["sets_lost"]
            player["game_diff"] = player["games_won"] - player["games_lost"]

        ranked_players = sorted(
            players,
            key=lambda item: (item["points"], item["set_diff"], item["game_diff"], item["games_won"]),
            reverse=True
        )
        top_players = ranked_players[:advancers_per_group]
        if len(top_players) < advancers_per_group:
            insufficient_groups.append(f"Bang {group_id}: co {len(top_players)}/{advancers_per_group}")
        group_tops.append(top_players)

    if insufficient_groups:
        raise ValueError("Mot so bang chua du VDV/cap co diem de chot Playoff: " + "; ".join(insufficient_groups))

    seeded_players = []
    for rank_index in range(advancers_per_group):
        for group in group_tops:
            if rank_index < len(group):
                seeded_players.append(group[rank_index])

    participant_count = len(seeded_players)
    if participant_count < 2:
        raise ValueError("Can it nhat 2 VDV/cap dau de tao Playoff.")

    db.query(Match).filter(
        Match.tournament_id == tournament_id,
        Match.tournament_category_id == category_id,
        Match.stage_type == "playoff"
    ).delete()
    db.flush()

    round_match_counts = []
    current_participants = participant_count
    while current_participants > 1:
        current_matches = math.ceil(current_participants / 2)
        round_match_counts.append(current_matches)
        current_participants = current_matches

    label_by_match_count = {
        16: "1/16",
        8: "1/8",
        4: "1/4",
        2: "1/2",
        1: "FINAL"
    }

    matches_by_round = {}
    for round_index, num_matches in enumerate(round_match_counts):
        matches_by_round[round_index] = []
        round_code = label_by_match_count.get(num_matches, f"R{num_matches * 2}")
        for idx in range(num_matches):
            new_match = Match(
                tournament_id=tournament_id,
                tournament_category_id=category_id,
                stage_type="playoff",
                round_code=round_code,
                match_no=idx + 1,
                status="pending",
                best_of_sets=3,
                elo_affected=True
            )
            db.add(new_match)
            matches_by_round[round_index].append(new_match)

    db.flush()

    incoming_counts = {}
    for round_index in range(0, len(round_match_counts) - 1):
        for match in matches_by_round[round_index]:
            for future_match in matches_by_round[round_index + 1]:
                current_incoming = incoming_counts.get(future_match.id, 0)
                if current_incoming < 2:
                    match.next_match_id = future_match.id
                    incoming_counts[future_match.id] = current_incoming + 1
                    break

    first_round_slots = seeded_players[:]
    target_slot_count = round_match_counts[0] * 2
    while len(first_round_slots) < target_slot_count:
        first_round_slots.append(None)

    first_round_pairs = []
    for index in range(round_match_counts[0]):
        first_round_pairs.append((first_round_slots[index], first_round_slots[target_slot_count - 1 - index]))

    for match, pair in zip(matches_by_round[0], first_round_pairs):
        side_a, side_b = pair
        match.side_a_registration_id = side_a["registration_id"] if side_a else None
        match.side_b_registration_id = side_b["registration_id"] if side_b else None

    db.commit()
    return {
        "message": "Da chot vong bang va tao so do Playoff thanh cong.",
        "qualified_count": participant_count,
        "first_round_matches": round_match_counts[0],
        "rounds": len(round_match_counts)
    }

def auto_update_tournament_statuses(db: Session):
    """Hàm chạy ngầm �Ồ quét và cập nhật trạng thái giải �ấu dựa trên thời gian thực tế."""
    today = datetime.utcnow().date()
    
    # 1. ChuyỒn từ 'open' sang 'ongoing' nếu �ã �ến ngày khai mạc
    open_tours = db.query(Tournament).filter(
        Tournament.status == "open",
        Tournament.start_date <= today
    ).all()
    for tour in open_tours:
        tour.status = "ongoing"
        
    # 2. ChuyỒn sang 'finished' nếu quá ngày kết thúc
    # Ch�0 quét các giải �ang m�x hoặc �ang di�&n ra mà �ã quá hạn
    past_tours = db.query(Tournament).filter(
        Tournament.status.in_(["open", "ongoing"]),
        Tournament.end_date < today
    ).all()
    for tour in past_tours:
        tour.status = "finished"
        
    db.commit()
    return len(open_tours) + len(past_tours)
